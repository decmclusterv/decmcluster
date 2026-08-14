import csv
import datetime
import io

import openpyxl

from ..models import Displacement


def find_missing_required_fields(file_obj):
    """
    Scans the uploaded file for missing required fields and returns a list of errors
    indicating row and column names.
    """
    file_name = getattr(file_obj, "name", "").lower()
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    rows = []
    if file_name.endswith(".csv"):
        content = file_obj.read()
        if isinstance(content, bytes):
            try:
                decoded = content.decode("utf-8")
            except UnicodeDecodeError:
                decoded = content.decode("latin-1")
        else:
            decoded = content
        csv_file = io.StringIO(decoded)
        reader = csv.reader(csv_file)
        rows = list(reader)
        
        header_index = 0
        for idx, row in enumerate(rows[:5]):
            if len(row) > 1 and "operation" in [cell.strip().lower() for cell in row]:
                header_index = idx
                break
        data_rows_start = header_index + 1
    else:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        rows = []
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 34)]
            rows.append(row_vals)
        data_rows_start = 1

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    required_cols = [
        (1, "Operation Code"),
        (2, "Operation"),
        (3, "Admin0 Name"),
        (4, "Admin0 Pcode"),
        (5, "Admin1 Name"),
        (6, "Admin1 Pcode"),
        (7, "Admin2 Name"),
        (8, "Admin2 Pcode"),
        (9, "Admin Level"),
        (10, "Num Present IDPs"),
        (11, "Reporting Date"),
        (12, "Reporting Year"),
        (13, "Reporting Month"),
        (14, "Round Number"),
        (15, "Displacement Reason"),
        (16, "Males Number"),
        (17, "Female Number"),
        (18, "Males Number 0-4"),
        (19, "Females Number 0-4"),
        (20, "Males Number 5-17"),
        (21, "Females Number 5-17"),
        (22, "Males Number 18-59"),
        (23, "Females Number 18-59"),
        (24, "Males Number 60+"),
        (25, "Females Number 60+"),
        (26, "Total Vul HHs"),
        (27, "IDP Origin Admin1 Name"),
        (28, "IDP Origin Admin1 Pcode"),
        (29, "Assessment Type"),
        (30, "Operation Status"),
        (31, "IDP Destination"),
        (32, "IDP Destination Admin1 Name"),
        (33, "IDP Destination Admin1 Pcode"),
    ]

    errors = []
    for row_idx, row in enumerate(rows):
        if row_idx < data_rows_start:
            continue

        if not any(str(val).strip() for val in row if val is not None):
            continue

        operation = row[1] if len(row) > 1 else None
        if not operation or not str(operation).strip():
            continue

        for col_idx, col_name in required_cols:
            val = None
            if len(row) >= col_idx:
                val = row[col_idx - 1]

            if val is None or str(val).strip() == "":
                errors.append(f"Row {row_idx + 1}: Column {col_idx} ({col_name}) is missing.")

    return errors


def to_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    try:
        return int(float(val_str))
    except ValueError:
        return None


def to_str(val):
    if val is None:
        return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    return val_str


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.date() if isinstance(val, datetime.datetime) else val
    if isinstance(val, (int, float)):
        try:
            # Excel serial date offset (Dec 30 1899)
            return datetime.date(1899, 12, 30) + datetime.timedelta(days=int(val))
        except Exception:
            return None
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            continue
    return None


def import_displacements_from_excel(file_path, uploaded_by=None, verified_by=None):
    """
    Parses Excel and directly creates Displacement records.
    """
    if hasattr(file_path, "read"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active

    created_count = 0

    # Headers at row 1, data starts at row 2
    for r in range(2, sheet.max_row + 1):
        operation_code = sheet.cell(row=r, column=1).value
        operation = sheet.cell(row=r, column=2).value
        admin0_name = sheet.cell(row=r, column=3).value
        admin0_pcode = sheet.cell(row=r, column=4).value
        admin1_name = sheet.cell(row=r, column=5).value
        admin1_pcode = sheet.cell(row=r, column=6).value
        admin2_name = sheet.cell(row=r, column=7).value
        admin2_pcode = sheet.cell(row=r, column=8).value
        admin_level = sheet.cell(row=r, column=9).value
        num_present_idps = sheet.cell(row=r, column=10).value
        reporting_date = sheet.cell(row=r, column=11).value
        reporting_year = sheet.cell(row=r, column=12).value
        reporting_month = sheet.cell(row=r, column=13).value
        round_number = sheet.cell(row=r, column=14).value
        displacement_reason = sheet.cell(row=r, column=15).value
        males_number = sheet.cell(row=r, column=16).value
        female_number = sheet.cell(row=r, column=17).value
        males_number_0_4 = sheet.cell(row=r, column=18).value
        females_number_0_4 = sheet.cell(row=r, column=19).value
        males_number_5_17 = sheet.cell(row=r, column=20).value
        females_number_5_17 = sheet.cell(row=r, column=21).value
        males_number_18_59 = sheet.cell(row=r, column=22).value
        females_number_18_59 = sheet.cell(row=r, column=23).value
        males_number_60_plus = sheet.cell(row=r, column=24).value
        females_number_60_plus = sheet.cell(row=r, column=25).value
        total_vul_hhs = sheet.cell(row=r, column=26).value
        idp_origin_admin1_name = sheet.cell(row=r, column=27).value
        idp_origin_admin1_pcode = sheet.cell(row=r, column=28).value
        assessment_type = sheet.cell(row=r, column=29).value
        operation_status = sheet.cell(row=r, column=30).value
        idp_destination = sheet.cell(row=r, column=31).value
        idp_destination_admin1_name = sheet.cell(row=r, column=32).value
        idp_destination_admin1_pcode = sheet.cell(row=r, column=33).value

        # Skip rows with no operation
        if not operation or not str(operation).strip():
            continue

        Displacement.objects.create(
            operation_code=to_str(operation_code),
            operation=to_str(operation),
            admin0_name=to_str(admin0_name),
            admin0_pcode=to_str(admin0_pcode),
            admin1_name=to_str(admin1_name),
            admin1_pcode=to_str(admin1_pcode),
            admin2_name=to_str(admin2_name),
            admin2_pcode=to_str(admin2_pcode),
            admin_level=to_int(admin_level),
            num_present_idps=to_int(num_present_idps),
            reporting_date=parse_date(reporting_date),
            reporting_year=to_int(reporting_year),
            reporting_month=to_int(reporting_month),
            round_number=to_int(round_number),
            displacement_reason=to_str(displacement_reason),
            males_number=to_int(males_number),
            female_number=to_int(female_number),
            males_number_0_4=to_int(males_number_0_4),
            females_number_0_4=to_int(females_number_0_4),
            males_number_5_17=to_int(males_number_5_17),
            females_number_5_17=to_int(females_number_5_17),
            males_number_18_59=to_int(males_number_18_59),
            females_number_18_59=to_int(females_number_18_59),
            males_number_60_plus=to_int(males_number_60_plus),
            females_number_60_plus=to_int(females_number_60_plus),
            total_vul_hhs=to_int(total_vul_hhs),
            idp_origin_admin1_name=to_str(idp_origin_admin1_name),
            idp_origin_admin1_pcode=to_str(idp_origin_admin1_pcode),
            assessment_type=to_str(assessment_type),
            operation_status=to_str(operation_status).lower()
            if operation_status
            else None,
            idp_destination=to_str(idp_destination),
            idp_destination_admin1_name=to_str(idp_destination_admin1_name),
            idp_destination_admin1_pcode=to_str(idp_destination_admin1_pcode),
            status=Displacement.StatusChoices.VERIFIED,
            uploaded_by=uploaded_by,
            verified_by=verified_by,
        )
        created_count += 1

    return created_count, 0


def import_displacements_from_csv(file_obj, uploaded_by=None, verified_by=None):
    """
    Parses CSV and directly creates Displacement records.
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        try:
            decoded_content = content.decode("utf-8")
        except UnicodeDecodeError:
            decoded_content = content.decode("latin-1")
    else:
        decoded_content = content

    csv_file = io.StringIO(decoded_content)
    reader = csv.reader(csv_file)
    rows = list(reader)

    # Find the header row (typically contains 'operation' or 'admin0name' in the first 5 rows)
    header_row_index = 0
    for idx, row in enumerate(rows[:5]):
        if len(row) > 1 and "operation" in [cell.strip().lower() for cell in row]:
            header_row_index = idx
            break

    data_rows = rows[header_row_index + 1 :]

    created_count = 0

    for row in data_rows:
        if not row or len(row) < 2:
            continue

        # Pad row with None to avoid index errors
        padded_row = row + [None] * (33 - len(row))

        operation_code = padded_row[0]
        operation = padded_row[1]
        admin0_name = padded_row[2]
        admin0_pcode = padded_row[3]
        admin1_name = padded_row[4]
        admin1_pcode = padded_row[5]
        admin2_name = padded_row[6]
        admin2_pcode = padded_row[7]
        admin_level = padded_row[8]
        num_present_idps = padded_row[9]
        reporting_date = padded_row[10]
        reporting_year = padded_row[11]
        reporting_month = padded_row[12]
        round_number = padded_row[13]
        displacement_reason = padded_row[14]
        males_number = padded_row[15]
        female_number = padded_row[16]
        males_number_0_4 = padded_row[17]
        females_number_0_4 = padded_row[18]
        males_number_5_17 = padded_row[19]
        females_number_5_17 = padded_row[20]
        males_number_18_59 = padded_row[21]
        females_number_18_59 = padded_row[22]
        males_number_60_plus = padded_row[23]
        females_number_60_plus = padded_row[24]
        total_vul_hhs = padded_row[25]
        idp_origin_admin1_name = padded_row[26]
        idp_origin_admin1_pcode = padded_row[27]
        assessment_type = padded_row[28]
        operation_status = padded_row[29]
        idp_destination = padded_row[30]
        idp_destination_admin1_name = padded_row[31]
        idp_destination_admin1_pcode = padded_row[32]

        if not operation or not str(operation).strip():
            continue

        Displacement.objects.create(
            operation_code=to_str(operation_code),
            operation=to_str(operation),
            admin0_name=to_str(admin0_name),
            admin0_pcode=to_str(admin0_pcode),
            admin1_name=to_str(admin1_name),
            admin1_pcode=to_str(admin1_pcode),
            admin2_name=to_str(admin2_name),
            admin2_pcode=to_str(admin2_pcode),
            admin_level=to_int(admin_level),
            num_present_idps=to_int(num_present_idps),
            reporting_date=parse_date(reporting_date),
            reporting_year=to_int(reporting_year),
            reporting_month=to_int(reporting_month),
            round_number=to_int(round_number),
            displacement_reason=to_str(displacement_reason),
            males_number=to_int(males_number),
            female_number=to_int(female_number),
            males_number_0_4=to_int(males_number_0_4),
            females_number_0_4=to_int(females_number_0_4),
            males_number_5_17=to_int(males_number_5_17),
            females_number_5_17=to_int(females_number_5_17),
            males_number_18_59=to_int(males_number_18_59),
            females_number_18_59=to_int(females_number_18_59),
            males_number_60_plus=to_int(males_number_60_plus),
            females_number_60_plus=to_int(females_number_60_plus),
            total_vul_hhs=to_int(total_vul_hhs),
            idp_origin_admin1_name=to_str(idp_origin_admin1_name),
            idp_origin_admin1_pcode=to_str(idp_origin_admin1_pcode),
            assessment_type=to_str(assessment_type),
            operation_status=to_str(operation_status).lower()
            if operation_status
            else None,
            idp_destination=to_str(idp_destination),
            idp_destination_admin1_name=to_str(idp_destination_admin1_name),
            idp_destination_admin1_pcode=to_str(idp_destination_admin1_pcode),
            status=Displacement.StatusChoices.VERIFIED,
            uploaded_by=uploaded_by,
            verified_by=verified_by,
        )
        created_count += 1

    return created_count, 0
