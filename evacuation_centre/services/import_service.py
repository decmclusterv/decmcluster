import csv
import io
import logging

import openpyxl

from ..models import EvacuationCentre

logger = logging.getLogger(__name__)


def to_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    val_str = str(val).strip()
    if not val_str or val_str.lower() in ("none", "null", "na", "n/a", "-"):
        return None
    try:
        return int(float(val_str))
    except ValueError:
        return None


def to_bool(val):
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    val_str = str(val).strip().lower()
    return val_str in ("yes", "true", "1")


def find_duplicate_compound_names(file_obj):
    """
    Scans an uploaded Excel/CSV file object and returns a list of duplicate compound names
    (duplicates within the file itself or matching existing compound names in DB).
    """
    file_name = getattr(file_obj, "name", "").lower()
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    raw_names = []
    if file_name.endswith(".csv"):
        content = file_obj.read()
        if isinstance(content, bytes):
            try:
                decoded = content.decode("utf-8")
            except UnicodeDecodeError:
                decoded = content.decode("latin-1")
        else:
            decoded = content
        csv_file = io.StringIO(decoded)
        reader = csv.reader(csv_file)
        rows = list(reader)

        header_index = 0
        for idx, row in enumerate(rows[:5]):
            if len(row) > 3 and (
                row[0].strip().lower() == "country"
                or row[1].strip().lower() == "organisation"
            ):
                header_index = idx
                break

        for row in rows[header_index + 1 :]:
            if len(row) > 3 and row[3] and str(row[3]).strip():
                raw_names.append(str(row[3]).strip())
    else:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        for r in range(4, sheet.max_row + 1):
            val = sheet.cell(row=r, column=4).value
            if val and str(val).strip():
                raw_names.append(str(val).strip())

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    duplicates = set()
    seen_in_file = set()

    existing_db_names = set(
        EvacuationCentre.objects.values_list("compound_name", flat=True)
    )
    existing_db_lower = {name.lower().strip() for name in existing_db_names if name}

    for name in raw_names:
        name_lower = name.lower().strip()
        if name_lower in seen_in_file or name_lower in existing_db_lower:
            duplicates.add(name)
        else:
            seen_in_file.add(name_lower)

    return list(duplicates)


def find_missing_required_fields(file_obj):
    """
    Scans the uploaded file for missing required fields and returns a list of errors
    indicating row and column names.
    """
    file_name = getattr(file_obj, "name", "").lower()
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    rows = []
    if file_name.endswith(".csv"):
        content = file_obj.read()
        if isinstance(content, bytes):
            try:
                decoded = content.decode("utf-8")
            except UnicodeDecodeError:
                decoded = content.decode("latin-1")
        else:
            decoded = content
        csv_file = io.StringIO(decoded)
        reader = csv.reader(csv_file)
        rows = list(reader)
        
        header_index = 0
        for idx, row in enumerate(rows[:5]):
            if len(row) > 3 and (
                row[0].strip().lower() == "country"
                or row[1].strip().lower() == "organisation"
            ):
                header_index = idx
                break
        data_rows_start = header_index + 1
    else:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        sheet = wb.active
        rows = []
        for r in range(1, sheet.max_row + 1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 40)]
            rows.append(row_vals)
        data_rows_start = 3

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    required_cols = [
        (1, "Country"),
        (2, "Organization"),
        (3, "Agency"),
        (4, "Compound Name"),
        (5, "Latitude"),
        (6, "Longitude"),
        (7, "Province"),
        (8, "Area Council"),
        (9, "Island"),
        (10, "Village"),
        (11, "Primary Contact"),
        (12, "Secondary Contact"),
        (13, "Compound Function"),
        (16, "Name of Outside Temporary Shelter"),
        (17, "Outside Temporary Shelter Capacity"),
        (20, "Electricity Source"),
        (21, "Drinking Water Source"),
        (22, "Washing Water Source"),
        (23, "Water Storage Capacity (Litres)"),
        (24, "No of Buildings"),
        (25, "No of Rooms"),
        (26, "Internal Building Evacuee Capacity"),
        (27, "Disaster Suitable For"),
        (28, "Engineering Certified Cyclone Rating"),
        (29, "Total Men's Toilet"),
        (30, "Total Women's Toilet"),
        (31, "Total Unisex Toilet"),
        (32, "Total Disability Access Toilet"),
        (33, "Total Men's Shower"),
        (34, "Total Women's Shower"),
        (35, "Total Unisex Shower"),
        (36, "Total Disability Access Shower"),
        (39, "Communication Back Up"),
    ]

    errors = []
    for row_idx, row in enumerate(rows):
        if row_idx < data_rows_start:
            continue

        if not any(str(val).strip() for val in row if val is not None):
            continue

        comp_name = row[3] if len(row) > 3 else None
        if not comp_name or not str(comp_name).strip():
            continue

        for col_idx, col_name in required_cols:
            val = None
            if len(row) >= col_idx:
                val = row[col_idx - 1]

            if val is None or str(val).strip() == "":
                errors.append(f"Row {row_idx + 1}: Column {col_idx} ({col_name}) is missing.")

    return errors


def import_evacuation_centres_from_excel(file_path):
    """
    Reads Evacuation Centres from an Excel sheet and creates them in the DB,
    skipping rows with duplicate compound names.
    """
    if hasattr(file_path, "read"):
        wb = openpyxl.load_workbook(file_path, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb.active

    existing_db_names = set(
        EvacuationCentre.objects.values_list("compound_name", flat=True)
    )
    seen_names = {name.lower().strip() for name in existing_db_names if name}
    created_count = 0
    skipped_count = 0

    # Headers start at row 3, data starts at row 4
    for r in range(4, sheet.max_row + 1):
        country = sheet.cell(row=r, column=1).value
        organization = sheet.cell(row=r, column=2).value
        agency = sheet.cell(row=r, column=3).value
        compound_name = sheet.cell(row=r, column=4).value
        latitude = sheet.cell(row=r, column=5).value
        longitude = sheet.cell(row=r, column=6).value
        province = sheet.cell(row=r, column=7).value
        area_council = sheet.cell(row=r, column=8).value
        island = sheet.cell(row=r, column=9).value
        village = sheet.cell(row=r, column=10).value
        primary_contact = sheet.cell(row=r, column=11).value
        secondary_contact = sheet.cell(row=r, column=12).value
        compound_function = sheet.cell(row=r, column=13).value
        is_ec_owner_approved = sheet.cell(row=r, column=14).value
        is_ec_govt_approved = sheet.cell(row=r, column=15).value
        name_of_outside_temporary_shelter = sheet.cell(row=r, column=16).value
        outside_temporary_shelter_capacity = sheet.cell(row=r, column=17).value
        first_aid_kit_availability = sheet.cell(row=r, column=18).value
        first_aid_trained_person = sheet.cell(row=r, column=19).value
        electricity_source = sheet.cell(row=r, column=20).value
        drinking_water_source = sheet.cell(row=r, column=21).value
        washing_water_source = sheet.cell(row=r, column=22).value
        water_storage_capacity_litres = sheet.cell(row=r, column=23).value
        no_of_buildings = sheet.cell(row=r, column=24).value
        no_of_rooms = sheet.cell(row=r, column=25).value
        internal_building_evacuee_capacity = sheet.cell(row=r, column=26).value
        disaster_suitable_for = sheet.cell(row=r, column=27).value
        enginerring_certified_cyclone_rating = sheet.cell(row=r, column=28).value
        total_mens_toilet = sheet.cell(row=r, column=29).value
        total_womens_toilet = sheet.cell(row=r, column=30).value
        total_unisex_toilet = sheet.cell(row=r, column=31).value
        total_disability_access_toilet = sheet.cell(row=r, column=32).value
        total_mens_shower = sheet.cell(row=r, column=33).value
        total_womens_shower = sheet.cell(row=r, column=34).value
        total_unisex_shower = sheet.cell(row=r, column=35).value
        total_disability_access_shower = sheet.cell(row=r, column=36).value
        kitchen_cooking_facilities = sheet.cell(row=r, column=37).value
        laundry_facilities = sheet.cell(row=r, column=38).value
        communication_back_up = sheet.cell(row=r, column=39).value

        # Skip rows with no compound name
        if not compound_name or not str(compound_name).strip():
            continue

        compound_name_clean = str(compound_name).strip()
        if compound_name_clean.lower() in seen_names:
            logger.warning(
                f"Duplicate data: compound name '{compound_name_clean}' already exists. Skipping row {r}."
            )
            skipped_count += 1
            continue

        seen_names.add(compound_name_clean.lower())

        # Coordinate parsing
        try:
            latitude_val = float(latitude) if latitude is not None else 0.0
            longitude_val = float(longitude) if longitude is not None else 0.0
        except ValueError:
            latitude_val = 0.0
            longitude_val = 0.0

        EvacuationCentre.objects.create(
            compound_name=compound_name_clean,
            latitude=latitude_val,
            longitude=longitude_val,
            country=str(country).strip() if country else "",
            organization=str(organization).strip() if organization else "",
            agency=str(agency).strip() if agency else "",
            province=str(province).strip() if province else "",
            area_council=str(area_council).strip() if area_council else "",
            island=str(island).strip() if island else None,
            village=str(village).strip() if village else None,
            primary_contact=str(primary_contact).strip() if primary_contact else None,
            secondary_contact=str(secondary_contact).strip()
            if secondary_contact
            else None,
            compound_function=str(compound_function).strip()
            if compound_function
            else None,
            is_ec_owner_approved=to_bool(is_ec_owner_approved),
            is_ec_govt_approved=to_bool(is_ec_govt_approved),
            name_of_outside_temporary_shelter=str(
                name_of_outside_temporary_shelter
            ).strip()
            if name_of_outside_temporary_shelter
            else None,
            outside_temporary_shelter_capacity=to_int(
                outside_temporary_shelter_capacity
            ),
            first_aid_kit_availability=to_bool(first_aid_kit_availability),
            first_aid_trained_person=to_bool(first_aid_trained_person),
            electricity_source=str(electricity_source).strip()
            if electricity_source
            else None,
            drinking_water_source=str(drinking_water_source).strip()
            if drinking_water_source
            else None,
            washing_water_source=str(washing_water_source).strip()
            if washing_water_source
            else None,
            water_storage_capacity_litres=to_int(water_storage_capacity_litres),
            no_of_buildings=to_int(no_of_buildings),
            no_of_rooms=to_int(no_of_rooms),
            internal_building_evacuee_capacity=to_int(
                internal_building_evacuee_capacity
            ),
            disaster_suitable_for=str(disaster_suitable_for).strip()
            if disaster_suitable_for
            else None,
            enginerring_certified_cyclone_rating=str(
                enginerring_certified_cyclone_rating
            ).strip()
            if enginerring_certified_cyclone_rating
            else None,
            total_mens_toilet=to_int(total_mens_toilet),
            total_womens_toilet=to_int(total_womens_toilet),
            total_unisex_toilet=to_int(total_unisex_toilet),
            total_disability_access_toilet=to_int(total_disability_access_toilet),
            total_mens_shower=to_int(total_mens_shower),
            total_womens_shower=to_int(total_womens_shower),
            total_unisex_shower=to_int(total_unisex_shower),
            total_disability_access_shower=to_int(total_disability_access_shower),
            kitchen_cooking_facilities=to_bool(kitchen_cooking_facilities),
            laundry_facilities=to_bool(laundry_facilities),
            communication_back_up=str(communication_back_up).strip()
            if communication_back_up
            else None,
        )
        created_count += 1

    return created_count, skipped_count


def import_evacuation_centres_from_csv(file_obj):
    """
    Reads Evacuation Centres from a CSV file and creates them in the DB,
    skipping rows with duplicate compound names.
    """
    content = file_obj.read()
    if isinstance(content, bytes):
        try:
            decoded_content = content.decode("utf-8")
        except UnicodeDecodeError:
            decoded_content = content.decode("latin-1")
    else:
        decoded_content = content

    csv_file = io.StringIO(decoded_content)
    reader = csv.reader(csv_file)
    rows = list(reader)

    # Find the header row (usually contains 'country' or 'organisation' in the first 5 rows)
    header_row_index = 0
    for idx, row in enumerate(rows[:5]):
        if len(row) > 3 and (
            row[0].strip().lower() == "country"
            or row[1].strip().lower() == "organisation"
        ):
            header_row_index = idx
            break

    data_rows = rows[header_row_index + 1 :]

    existing_db_names = set(
        EvacuationCentre.objects.values_list("compound_name", flat=True)
    )
    seen_names = {name.lower().strip() for name in existing_db_names if name}
    created_count = 0
    skipped_count = 0

    for row in data_rows:
        if not row or len(row) < 4:
            continue

        # Pad with None to avoid IndexError if there are fewer columns
        padded_row = row + [None] * (39 - len(row))

        country = padded_row[0]
        organization = padded_row[1]
        agency = padded_row[2]
        compound_name = padded_row[3]
        latitude = padded_row[4]
        longitude = padded_row[5]
        province = padded_row[6]
        area_council = padded_row[7]
        island = padded_row[8]
        village = padded_row[9]
        primary_contact = padded_row[10]
        secondary_contact = padded_row[11]
        compound_function = padded_row[12]
        is_ec_owner_approved = padded_row[13]
        is_ec_govt_approved = padded_row[14]
        name_of_outside_temporary_shelter = padded_row[15]
        outside_temporary_shelter_capacity = padded_row[16]
        first_aid_kit_availability = padded_row[17]
        first_aid_trained_person = padded_row[18]
        electricity_source = padded_row[19]
        drinking_water_source = padded_row[20]
        washing_water_source = padded_row[21]
        water_storage_capacity_litres = padded_row[22]
        no_of_buildings = padded_row[23]
        no_of_rooms = padded_row[24]
        internal_building_evacuee_capacity = padded_row[25]
        disaster_suitable_for = padded_row[26]
        enginerring_certified_cyclone_rating = padded_row[27]
        total_mens_toilet = padded_row[28]
        total_womens_toilet = padded_row[29]
        total_unisex_toilet = padded_row[30]
        total_disability_access_toilet = padded_row[31]
        total_mens_shower = padded_row[32]
        total_womens_shower = padded_row[33]
        total_unisex_shower = padded_row[34]
        total_disability_access_shower = padded_row[35]
        kitchen_cooking_facilities = padded_row[36]
        laundry_facilities = padded_row[37]
        communication_back_up = padded_row[38]

        # Skip rows with no compound name
        if not compound_name or not str(compound_name).strip():
            continue

        compound_name_clean = str(compound_name).strip()
        if compound_name_clean.lower() in seen_names:
            logger.warning(
                f"Duplicate data: compound name '{compound_name_clean}' already exists. Skipping."
            )
            skipped_count += 1
            continue

        seen_names.add(compound_name_clean.lower())

        # Coordinate parsing
        try:
            latitude_val = (
                float(latitude)
                if latitude is not None and str(latitude).strip() != ""
                else 0.0
            )
            longitude_val = (
                float(longitude)
                if longitude is not None and str(longitude).strip() != ""
                else 0.0
            )
        except ValueError:
            latitude_val = 0.0
            longitude_val = 0.0

        EvacuationCentre.objects.create(
            compound_name=compound_name_clean,
            latitude=latitude_val,
            longitude=longitude_val,
            country=str(country).strip() if country else "",
            organization=str(organization).strip() if organization else "",
            agency=str(agency).strip() if agency else "",
            province=str(province).strip() if province else "",
            area_council=str(area_council).strip() if area_council else "",
            island=str(island).strip() if island else None,
            village=str(village).strip() if village else None,
            primary_contact=str(primary_contact).strip() if primary_contact else None,
            secondary_contact=str(secondary_contact).strip()
            if secondary_contact
            else None,
            compound_function=str(compound_function).strip()
            if compound_function
            else None,
            is_ec_owner_approved=to_bool(is_ec_owner_approved),
            is_ec_govt_approved=to_bool(is_ec_govt_approved),
            name_of_outside_temporary_shelter=str(
                name_of_outside_temporary_shelter
            ).strip()
            if name_of_outside_temporary_shelter
            else None,
            outside_temporary_shelter_capacity=to_int(
                outside_temporary_shelter_capacity
            ),
            first_aid_kit_availability=to_bool(first_aid_kit_availability),
            first_aid_trained_person=to_bool(first_aid_trained_person),
            electricity_source=str(electricity_source).strip()
            if electricity_source
            else None,
            drinking_water_source=str(drinking_water_source).strip()
            if drinking_water_source
            else None,
            washing_water_source=str(washing_water_source).strip()
            if washing_water_source
            else None,
            water_storage_capacity_litres=to_int(water_storage_capacity_litres),
            no_of_buildings=to_int(no_of_buildings),
            no_of_rooms=to_int(no_of_rooms),
            internal_building_evacuee_capacity=to_int(
                internal_building_evacuee_capacity
            ),
            disaster_suitable_for=str(disaster_suitable_for).strip()
            if disaster_suitable_for
            else None,
            enginerring_certified_cyclone_rating=str(
                enginerring_certified_cyclone_rating
            ).strip()
            if enginerring_certified_cyclone_rating
            else None,
            total_mens_toilet=to_int(total_mens_toilet),
            total_womens_toilet=to_int(total_womens_toilet),
            total_unisex_toilet=to_int(total_unisex_toilet),
            total_disability_access_toilet=to_int(total_disability_access_toilet),
            total_mens_shower=to_int(total_mens_shower),
            total_womens_shower=to_int(total_womens_shower),
            total_unisex_shower=to_int(total_unisex_shower),
            total_disability_access_shower=to_int(total_disability_access_shower),
            kitchen_cooking_facilities=to_bool(kitchen_cooking_facilities),
            laundry_facilities=to_bool(laundry_facilities),
            communication_back_up=str(communication_back_up).strip()
            if communication_back_up
            else None,
        )
        created_count += 1

    return created_count, skipped_count
